#!/usr/bin/env python3
"""
ATHLYNX CRM Excel Export Script
Exports signup data to Excel format ready for Microsoft Copilot analysis
"""

import os
import sys
from datetime import datetime
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList

# Database connection
DATABASE_URL = os.environ.get('DATABASE_URL', '')

def parse_database_url(url):
    """Parse DATABASE_URL into connection parameters"""
    # Format: mysql://user:pass@host:port/database
    import re
    pattern = r'mysql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)'
    match = re.match(pattern, url)
    if match:
        return {
            'user': match.group(1),
            'password': match.group(2),
            'host': match.group(3),
            'port': int(match.group(4)),
            'database': match.group(5)
        }
    return None

def create_excel_report():
    """Create comprehensive Excel report for Microsoft Copilot"""
    
    # Connect to database
    db_config = parse_database_url(DATABASE_URL)
    if not db_config:
        print("Error: Invalid DATABASE_URL")
        sys.exit(1)
    
    conn = mysql.connector.connect(**db_config, ssl_disabled=False)
    cursor = conn.cursor(dictionary=True)
    
    # Create workbook
    wb = Workbook()
    
    # ==================== SHEET 1: SIGNUPS ====================
    ws_signups = wb.active
    ws_signups.title = "Signups"
    
    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Signup #", "Timestamp", "Full Name", "Email", "Phone", "Role", "Sport",
        "IP Address", "Browser", "Device", "OS", "Country", "City",
        "Referral Source", "UTM Source", "UTM Medium", "UTM Campaign",
        "Signup Type", "Converted", "Paying", "Lifetime Value ($)"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_signups.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Fetch signup data
    cursor.execute("""
        SELECT * FROM signup_analytics 
        ORDER BY signupNumber ASC
    """)
    signups = cursor.fetchall()
    
    # Data rows
    for row_num, signup in enumerate(signups, 2):
        data = [
            signup.get('signupNumber', ''),
            signup.get('createdAt', '').strftime('%Y-%m-%d %H:%M:%S') if signup.get('createdAt') else '',
            signup.get('fullName', ''),
            signup.get('email', ''),
            signup.get('phone', ''),
            signup.get('role', ''),
            signup.get('sport', ''),
            signup.get('ipAddress', ''),
            signup.get('browser', ''),
            signup.get('device', ''),
            signup.get('os', ''),
            signup.get('country', ''),
            signup.get('city', ''),
            signup.get('referralSource', ''),
            signup.get('utmSource', ''),
            signup.get('utmMedium', ''),
            signup.get('utmCampaign', ''),
            signup.get('signupType', ''),
            'Yes' if signup.get('isConverted') else 'No',
            'Yes' if signup.get('isPaying') else 'No',
            float(signup.get('lifetimeValue', 0) or 0)
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws_signups.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col == 1:  # Signup number
                cell.font = Font(bold=True, color="0066CC")
    
    # Auto-fit columns
    for col in ws_signups.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_signups.column_dimensions[column].width = adjusted_width
    
    # ==================== SHEET 2: SUMMARY STATS ====================
    ws_stats = wb.create_sheet("Summary Stats")
    
    # Calculate stats
    cursor.execute("SELECT COUNT(*) as total FROM signup_analytics")
    total_signups = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as today FROM signup_analytics WHERE DATE(createdAt) = CURDATE()")
    today_signups = cursor.fetchone()['today']
    
    cursor.execute("SELECT COUNT(*) as converted FROM signup_analytics WHERE isConverted = 1")
    converted = cursor.fetchone()['converted']
    
    cursor.execute("SELECT COUNT(*) as paying FROM signup_analytics WHERE isPaying = 1")
    paying = cursor.fetchone()['paying']
    
    cursor.execute("SELECT COALESCE(SUM(lifetimeValue), 0) as revenue FROM signup_analytics")
    revenue = float(cursor.fetchone()['revenue'])
    
    # Stats table
    stats = [
        ("ATHLYNX CRM REPORT", ""),
        ("Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ("", ""),
        ("KEY METRICS", ""),
        ("Total Signups", total_signups),
        ("Today's Signups", today_signups),
        ("Converted Users", converted),
        ("Paying Customers", paying),
        ("Total Revenue", f"${revenue:,.2f}"),
        ("Conversion Rate", f"{(converted/total_signups*100) if total_signups > 0 else 0:.2f}%"),
        ("", ""),
        ("COPILOT ANALYSIS PROMPTS", ""),
        ("", ""),
        ("Try asking Copilot:", ""),
        ("1.", "What are the top 5 sports by signup count?"),
        ("2.", "Show me the signup trend over time"),
        ("3.", "Which referral sources are most effective?"),
        ("4.", "What's the conversion rate by role type?"),
        ("5.", "Identify patterns in paying customers"),
        ("6.", "Create a forecast for next month's signups"),
    ]
    
    for row_num, (label, value) in enumerate(stats, 1):
        ws_stats.cell(row=row_num, column=1, value=label)
        ws_stats.cell(row=row_num, column=2, value=value)
        if label in ["ATHLYNX CRM REPORT", "KEY METRICS", "COPILOT ANALYSIS PROMPTS"]:
            ws_stats.cell(row=row_num, column=1).font = Font(bold=True, size=14, color="0066CC")
    
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 40
    
    # ==================== SHEET 3: BY ROLE ====================
    ws_role = wb.create_sheet("By Role")
    
    cursor.execute("""
        SELECT role, COUNT(*) as count, 
               SUM(CASE WHEN isConverted = 1 THEN 1 ELSE 0 END) as converted,
               SUM(CASE WHEN isPaying = 1 THEN 1 ELSE 0 END) as paying,
               COALESCE(SUM(lifetimeValue), 0) as revenue
        FROM signup_analytics 
        GROUP BY role
        ORDER BY count DESC
    """)
    role_data = cursor.fetchall()
    
    role_headers = ["Role", "Total Signups", "Converted", "Paying", "Revenue ($)"]
    for col, header in enumerate(role_headers, 1):
        cell = ws_role.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_num, role in enumerate(role_data, 2):
        ws_role.cell(row=row_num, column=1, value=role['role'] or 'Unknown')
        ws_role.cell(row=row_num, column=2, value=role['count'])
        ws_role.cell(row=row_num, column=3, value=role['converted'])
        ws_role.cell(row=row_num, column=4, value=role['paying'])
        ws_role.cell(row=row_num, column=5, value=float(role['revenue']))
    
    # ==================== SHEET 4: BY SPORT ====================
    ws_sport = wb.create_sheet("By Sport")
    
    cursor.execute("""
        SELECT sport, COUNT(*) as count
        FROM signup_analytics 
        WHERE sport IS NOT NULL AND sport != ''
        GROUP BY sport
        ORDER BY count DESC
    """)
    sport_data = cursor.fetchall()
    
    sport_headers = ["Sport", "Total Signups"]
    for col, header in enumerate(sport_headers, 1):
        cell = ws_sport.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_num, sport in enumerate(sport_data, 2):
        ws_sport.cell(row=row_num, column=1, value=sport['sport'])
        ws_sport.cell(row=row_num, column=2, value=sport['count'])
    
    # ==================== SHEET 5: DAILY TREND ====================
    ws_trend = wb.create_sheet("Daily Trend")
    
    cursor.execute("""
        SELECT DATE(createdAt) as date, COUNT(*) as signups
        FROM signup_analytics 
        GROUP BY DATE(createdAt)
        ORDER BY date ASC
    """)
    trend_data = cursor.fetchall()
    
    trend_headers = ["Date", "Signups", "Cumulative"]
    for col, header in enumerate(trend_headers, 1):
        cell = ws_trend.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    cumulative = 0
    for row_num, day in enumerate(trend_data, 2):
        cumulative += day['signups']
        ws_trend.cell(row=row_num, column=1, value=day['date'].strftime('%Y-%m-%d') if day['date'] else '')
        ws_trend.cell(row=row_num, column=2, value=day['signups'])
        ws_trend.cell(row=row_num, column=3, value=cumulative)
    
    # ==================== SHEET 6: MILESTONES ====================
    ws_milestones = wb.create_sheet("Milestones")
    
    cursor.execute("SELECT * FROM milestones ORDER BY targetValue ASC")
    milestones = cursor.fetchall()
    
    milestone_headers = ["Milestone", "Type", "Target", "Current", "Progress %", "Achieved", "Achieved At"]
    for col, header in enumerate(milestone_headers, 1):
        cell = ws_milestones.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    for row_num, milestone in enumerate(milestones, 2):
        target = milestone.get('targetValue', 1)
        current = milestone.get('currentValue', 0)
        progress = min((current / target * 100) if target > 0 else 0, 100)
        
        ws_milestones.cell(row=row_num, column=1, value=milestone.get('milestoneName', ''))
        ws_milestones.cell(row=row_num, column=2, value=milestone.get('milestoneType', ''))
        ws_milestones.cell(row=row_num, column=3, value=target)
        ws_milestones.cell(row=row_num, column=4, value=current)
        ws_milestones.cell(row=row_num, column=5, value=f"{progress:.1f}%")
        ws_milestones.cell(row=row_num, column=6, value='Yes' if milestone.get('isAchieved') else 'No')
        ws_milestones.cell(row=row_num, column=7, value=milestone.get('achievedAt', ''))
        
        # Highlight achieved milestones
        if milestone.get('isAchieved'):
            for col in range(1, 8):
                ws_milestones.cell(row=row_num, column=col).fill = PatternFill(
                    start_color="90EE90", end_color="90EE90", fill_type="solid"
                )
    
    # Save workbook
    output_path = f"/home/ubuntu/athlynx-perfect-storm/exports/ATHLYNX_CRM_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ Excel report saved to: {output_path}")
    print(f"\n📊 Report Summary:")
    print(f"   Total Signups: {total_signups}")
    print(f"   Converted: {converted}")
    print(f"   Paying: {paying}")
    print(f"   Revenue: ${revenue:,.2f}")
    print(f"\n💡 Open in Excel and use Microsoft Copilot for AI-powered analysis!")
    
    cursor.close()
    conn.close()
    
    return output_path

if __name__ == "__main__":
    create_excel_report()
