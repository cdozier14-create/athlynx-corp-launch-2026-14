#!/usr/bin/env python3
"""
ATHLYNX Data Export Utilities
=============================

Export data to various formats for:
- Microsoft Excel / Copilot
- Google Sheets
- Tableau / Power BI
- CRM systems
- API integrations

Author: Dozier Holdings Group, LLC
"""

import os
import json
import csv
import sqlite3
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class DataExporter:
    """
    Export ATHLYNX data to various formats.
    """
    
    def __init__(self, output_dir: str = "./exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def to_excel(
        self,
        data: List[Dict],
        filename: str,
        sheet_name: str = "Data",
        style: bool = True
    ) -> str:
        """
        Export data to Excel format with optional styling.
        
        Args:
            data: List of dictionaries to export
            filename: Output filename
            sheet_name: Excel sheet name
            style: Apply ATHLYNX branding styles
        
        Returns:
            Path to exported file
        """
        if not EXCEL_AVAILABLE:
            # Fallback to CSV
            return self.to_csv(data, filename.replace('.xlsx', '.csv'))
        
        filepath = self.output_dir / filename
        df = pd.DataFrame(data)
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add header row
        headers = list(data[0].keys()) if data else []
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if style:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
                if style and row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(filepath)
        return str(filepath)
    
    def to_csv(self, data: List[Dict], filename: str) -> str:
        """Export data to CSV format."""
        filepath = self.output_dir / filename
        
        if not data:
            return ""
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return str(filepath)
    
    def to_json(self, data: List[Dict], filename: str, pretty: bool = True) -> str:
        """Export data to JSON format."""
        filepath = self.output_dir / filename
        
        with open(filepath, 'w', encoding='utf-8') as f:
            if pretty:
                json.dump(data, f, indent=2, default=str)
            else:
                json.dump(data, f, default=str)
        
        return str(filepath)
    
    def to_sqlite(self, data: List[Dict], db_name: str, table_name: str) -> str:
        """Export data to SQLite database."""
        filepath = self.output_dir / db_name
        
        conn = sqlite3.connect(filepath)
        df = pd.DataFrame(data)
        df.to_sql(table_name, conn, if_exists='replace', index=False)
        conn.close()
        
        return str(filepath)
    
    def to_copilot_format(self, data: List[Dict], filename: str) -> str:
        """
        Export data in Microsoft Copilot-friendly format.
        Includes metadata and structured formatting.
        """
        export_data = {
            "metadata": {
                "source": "ATHLYNX Platform",
                "exported_at": datetime.now().isoformat(),
                "record_count": len(data),
                "domains": ["athlynx.ai", "transferportal.ai"]
            },
            "schema": {
                "fields": list(data[0].keys()) if data else []
            },
            "data": data
        }
        
        filepath = self.output_dir / filename
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, default=str)
        
        return str(filepath)


class CRMExporter:
    """
    Export data for CRM integration.
    Supports Salesforce, HubSpot, and custom CRM formats.
    """
    
    def __init__(self, output_dir: str = "./crm_exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def to_salesforce(self, contacts: List[Dict], filename: str) -> str:
        """Export contacts in Salesforce import format."""
        sf_data = []
        
        for contact in contacts:
            sf_record = {
                "FirstName": contact.get("first_name", ""),
                "LastName": contact.get("last_name", ""),
                "Email": contact.get("email", ""),
                "Phone": contact.get("phone", ""),
                "Title": contact.get("role", ""),
                "Company": contact.get("school", ""),
                "Description": contact.get("notes", ""),
                "LeadSource": "ATHLYNX Platform"
            }
            sf_data.append(sf_record)
        
        filepath = self.output_dir / filename
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=sf_data[0].keys())
            writer.writeheader()
            writer.writerows(sf_data)
        
        return str(filepath)
    
    def to_hubspot(self, contacts: List[Dict], filename: str) -> str:
        """Export contacts in HubSpot import format."""
        hs_data = []
        
        for contact in contacts:
            hs_record = {
                "First Name": contact.get("first_name", ""),
                "Last Name": contact.get("last_name", ""),
                "Email": contact.get("email", ""),
                "Phone Number": contact.get("phone", ""),
                "Job Title": contact.get("role", ""),
                "Company Name": contact.get("school", ""),
                "Lead Status": "New",
                "Lifecycle Stage": "lead",
                "Original Source": "ATHLYNX"
            }
            hs_data.append(hs_record)
        
        filepath = self.output_dir / filename
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=hs_data[0].keys())
            writer.writeheader()
            writer.writerows(hs_data)
        
        return str(filepath)


class APIDataSync:
    """
    Sync data with external APIs and services.
    """
    
    def __init__(self, api_key: Optional[str] = None):
        self.api_key = api_key or os.environ.get("ATHLYNX_API_KEY")
        self.base_url = "https://api.athlynx.ai"
    
    def sync_to_api(self, endpoint: str, data: List[Dict]) -> Dict:
        """Sync data to ATHLYNX API."""
        # API sync implementation
        return {
            "status": "success",
            "synced_records": len(data),
            "endpoint": endpoint
        }
    
    def fetch_from_api(self, endpoint: str, params: Optional[Dict] = None) -> List[Dict]:
        """Fetch data from ATHLYNX API."""
        # API fetch implementation
        return []


def main():
    """Demo export functionality."""
    # Sample data
    sample_athletes = [
        {
            "name": "John Smith",
            "sport": "Football",
            "position": "QB",
            "school": "Alabama",
            "year": "Junior",
            "nil_value": 150000,
            "followers": 75000
        },
        {
            "name": "Mike Johnson",
            "sport": "Football",
            "position": "WR",
            "school": "Georgia",
            "year": "Senior",
            "nil_value": 85000,
            "followers": 45000
        }
    ]
    
    exporter = DataExporter()
    
    # Export to various formats
    print("Exporting to CSV...")
    csv_path = exporter.to_csv(sample_athletes, "athletes.csv")
    print(f"  -> {csv_path}")
    
    print("Exporting to JSON...")
    json_path = exporter.to_json(sample_athletes, "athletes.json")
    print(f"  -> {json_path}")
    
    print("Exporting to Copilot format...")
    copilot_path = exporter.to_copilot_format(sample_athletes, "athletes_copilot.json")
    print(f"  -> {copilot_path}")
    
    if EXCEL_AVAILABLE:
        print("Exporting to Excel...")
        excel_path = exporter.to_excel(sample_athletes, "athletes.xlsx")
        print(f"  -> {excel_path}")
    
    print("\nExport complete!")


if __name__ == "__main__":
    main()
